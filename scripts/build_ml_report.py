#!/usr/bin/env python3
"""
build_ml_report.py

Genera el .xlsx de 3 hojas (Ventas / Picking del dia / Pedido a taller)
a partir del Excel exportado de Mercado Libre (SRC). Mantiene control
de ventas ya contadas (CSV) y aplica las reglas del catálogo/excepciones.

Uso:
  python3 build_ml_report.py src.xlsx out.xlsx ~/Desktop/Gymaholic - Control Ventas ML/ventas_ya_contadas.csv
"""
import openpyxl, re, sys, os, csv, datetime
from collections import defaultdict
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------
# Config / Catálogo / Excepciones (mantener sincronizado con prompts.txt)
# -----------------------
CATALOGO = [
    ("Fitbell Balon Medicinal Crossfit Entrenamiento 4kg", "FITBELL4K", 74),
    ("Balon Medicinal Medball Fitbell Crossfit 5 Kg", "FITBELL5K", 75),
    ("Balon Medicinal Medball Fitbell Crossfit 6 Kg", "FITBELL6K", 90),
    ("Balon Medicinal Medball Fitbell Crossfit 8 Kg", "FITBELL8K", 110),
    ("Balon Medicinal Medball Crossfit 10 Kg", "MEDBALL10K (SIN AG)", 105),
    ("Balón Medicinal 5 Kg Entrenamiento Funcional Reforzado Gym", "BMED5K (SINAG)", 78),
    ("Sandbag Costal De Arena 10 Kg", "SANDBAG10K", 225),
    ("Sandbag Costal 10 Kg Entrenamiento", "SANDBAG10K", 225),
    ("Sandbag Costal De Arena 15 Kg", "SANDBAG15K", 225),
    ("Sandbag Costal De Arena 20 Kg", "SANDBAG20K", 225),
    ("Sandbag Costal De Arena 25 Kg", "SANDBAG25K", 240),
    ("Sandbag Costal De Arena 30 Kg", "SANDBAG30K", 240),
    ("Balon De Azote 6 Kg", "BAZOTE6K", 170),
    ("Balon De Azote 10 Kg", "BAZOTE10K", 230),
    ("Balon De Azote 15 Kg", "BAZOTE15K", 280),
    ("Balon De Azote 16 Kg", "BAZOTE16K", 300),
    ("Costal Búlgaro 5 Kg", "BULG5K", 225),
    ("Costal Búlgaro 10 Kg", "BULG10K", 225),
    ("Costal Búlgaro 15 Kg", "BULG15K", 225),
    ("Costal Búlgaro 20 Kg", "BULG20K", 225),
    ("Costal Búlgaro 25 Kg", "BULG25K", 240),
    ("Costal Búlgaro 30 Kg", "BULG30K", 240),
    ("Set Balones Medicinales 2, 4, 6, 8 Y 10", "SETFB", 550),
    ("Polainas Ajustables 10 Kg", "PAJ10K", 280),
    ("Polainas Peso Ajustable 2 A 10 Kg Por Lado", "PAJ20K", 480),
    ("Set 4 Balones Azote 4 + 6 + 8 + 10 Kg Reforzado Slam Ball", "SETBAZOTE", 750),
    ("Banco Sentadilla Sissy", "SISSY", 1800),
    ("Cinturón / Faja, Soporte, Pesas, Gym, Lumbar", "CINTPIEL", 250),
    ("Cinturón Faja Pesas Gymaholicmx Hecho En Mex", "CINTECO", 50),
    ("Chaleco Peso Ajustable 2 A 20 Kg Entrenamiento Hyrox Fitness", "CHALECO", 600),
    ("Bandas De Oclusión Para Crecimiento De Gluteo", "BO", 110),
]

# Excepciones por # de publicación
PUBLICACION_OVERRIDES = {"MLM3113854054": ("FITBELL10K", 130)}

# Proveedores externos (excluir de Pedido a taller)
PROVEEDORES_EXTERNOS = {"PAJ10K","PAJ20K","SISSY","CINTPIEL","CINTECO","CHALECO","BO"}

# KITFUNC componentes (desglose obligatorio)
KITFUNC_COMPONENTES = ["BULG20K", "SANDBAG10K", "FITBELL5K"]

# Códigos que llevan etiqueta EMPLAYADO (singular)
EMPLAYADOS_CODES = {"SETFB", "SETFITBELL", "SETBAZOTE"}

# Normalización de colores
COLOR_MAP = {"rosa":"ROSA","negro":"NEGRO","rojo":"ROJO","azul":"AZUL","naranja":"NARANJA","amarillo":"AMARILLO","morado":"MORADO","violeta":"MORADO","verde militar":"VERDE MILITAR","gris":"GRIS","lila":"LILA"}

# Estilos
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_BOLD = Font(name='Calibri', size=10, bold=True)
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# -----------------------
# Helpers
# -----------------------
def match_catalogo(titulo):
    for keyword, code, cost in CATALOGO:
        if keyword.lower() in (titulo or "").lower():
            return code, cost
    return None, None

def extract_color(variante):
    if not variante:
        return ""
    color_m = re.search(r'Color\s*:\s*([^|]+)', variante, re.IGNORECASE)
    talla_m = re.search(r'Talla\s*:\s*([^|]+)', variante, re.IGNORECASE)
    parts = []
    if color_m:
        val = color_m.group(1).strip().lower()
        parts.append(COLOR_MAP.get(val, val.upper()))
    elif not talla_m:
        m = re.search(r':\s*(.+)', variante)
        val = (m.group(1) if m else variante).strip().lower()
        parts.append(COLOR_MAP.get(val, val.upper()))
    if talla_m:
        parts.append(talla_m.group(1).strip().upper())
    return " ".join(parts)

def format_producto(code, color):
    producto = f"{code} {color}".strip()
    if code in EMPLAYADOS_CODES:
        producto += " (EMPLAYADO)"
    return producto

def setcell(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_BOLD if bold else FONT_NORMAL
    c.border = BORDER
    return c

def load_control(path):
    seen = set()
    if path and os.path.exists(path):
        with open(path, newline='') as f:
            for row in csv.reader(f):
                if row and row[0] != 'venta_id':
                    seen.add(row[0].strip())
    return seen

def append_control(path, nuevas, fecha):
    if not path:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    is_new_file = not os.path.exists(path)
    with open(path, 'a', newline='') as f:
        w = csv.writer(f)
        if is_new_file:
            w.writerow(['venta_id', 'primera_fecha_contada_en_ventas'])
        for vid in nuevas:
            w.writerow([vid, fecha])

# -----------------------
# Main processing
# -----------------------
def build_report(src_path, out_path, control_csv=None):
    ya_contadas = load_control(control_csv)

    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src.active

    header_row = None
    for r in range(1, 15):
        vals = [c.value for c in ws_src[r]]
        if vals and vals[0] and str(vals[0]).strip().lower() in ('# de venta','# de venta'.lower()):
            header_row = r
            break
    if header_row is None:
        raise SystemExit("No se encontro la fila de encabezados '# de venta'")

    headers = [c.value for c in ws_src[header_row]]
    idx = {}
    for i, h in enumerate(headers):
        if h not in idx:
            idx[h] = i

    ventas_rows = []
    picking = defaultdict(int)
    taller = defaultdict(int)
    sin_match = []
    nuevas_vids = []

    for row in ws_src.iter_rows(min_row=header_row+1, values_only=True):
        vid = row[idx['# de venta']]
        if vid is None or str(vid).strip() == '':
            continue
        vid_str = str(vid).strip()
        if vid_str in []:  # VENTAS_DUPLICADAS placeholder
            continue
        total_raw = row[idx['Total (MXN)']] if idx.get('Total (MXN)') is not None else None
        try:
            total_num = float(total_raw) if total_raw not in (None,"") else None
        except Exception:
            total_num = None

        es_nueva = vid_str not in ya_contadas

        # Ventas overrides (si aplica) - implement if you maintain VENTA_OVERRIDES
        # Example omitted for brevity

        unidades = int(row[idx['Unidades']] or 1)
        titulo = row[idx['Título de la publicación']] or ''
        variante = row[idx['Variante']] or ''
        publicacion = str(row[idx['# de publicación']]).strip() if idx.get('# de publicación') is not None else ''

        if publicacion in PUBLICACION_OVERRIDES:
            code, unit_cost = PUBLICACION_OVERRIDES[publicacion]
        else:
            code, unit_cost = match_catalogo(titulo)

        color = extract_color(variante)
        if code is None:
            sin_match.append((vid_str, titulo, variante))
            continue

        producto = format_producto(code, color)
        costo_total = (unit_cost or 0) * unidades

        picking[producto] += unidades
        if code not in PROVEEDORES_EXTERNOS:
            taller[producto] += unidades

        if es_nueva:
            prod_str = f"{unidades} {producto}" if unidades > 1 else producto
            ventas_rows.append((vid_str, prod_str, costo_total, total_num))
            nuevas_vids.append(vid_str)

    if sin_match:
        print("!!! PRODUCTOS SIN MATCH EN CATALOGO (pregunta al usuario codigo y costo):")
        for vid, t, v in sin_match:
            print(f"  venta {vid}: {t!r} / {v!r}")
        raise SystemExit(1)

    # Build workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ventas"

    headers_def = ['VENTA','PRODUCTO','','','COSTO','','TOTAL RECIBIDO']
    for col, title in enumerate(headers_def, start=1):
        setcell(ws1, 1, col, title, bold=True)
    for col, width in {'A':20.0,'B':42.0,'C':10.0,'D':4.0,'E':14.0,'F':4.0,'G':16.0}.items():
        ws1.column_dimensions[col].width = width

    r = 2
    tot_costo = 0.0
    tot_total = 0.0
    for vid, prod_str, costo_total, total in ventas_rows:
        # escribir venta_id COMO TEXTO y forzar formato de celda '@' para evitar truncado
        c_vid = setcell(ws1, r, 1, str(vid))
        c_vid.number_format = '@'
        setcell(ws1, r, 2, prod_str.upper())
        setcell(ws1, r, 3, 'FELIX')
        setcell(ws1, r, 4, '')
        setcell(ws1, r, 5, costo_total)
        setcell(ws1, r, 6, '')
        if total is None:
            setcell(ws1, r, 7, "")
        else:
            # escribir total con try/except y redondeo
            try:
                setcell(ws1, r, 7, round(float(total), 2))
            except Exception:
                setcell(ws1, r, 7, total)
        tot_costo += costo_total
        if isinstance(total, (int,float)):
            tot_total += float(total)
        r += 1

    r += 1
    setcell(ws1, r, 1, 'TOTAL', bold=True)
    setcell(ws1, r, 5, tot_costo, bold=True)
    setcell(ws1, r, 7, round(tot_total, 2), bold=True)

    def write_agg_sheet(title, data):
        ws = wb.create_sheet(title)
        setcell(ws, 1, 1, 'CANTIDAD', bold=True)
        setcell(ws, 1, 2, 'PRODUCTO', bold=True)
        ws.column_dimensions['A'].width = 12.0
        ws.column_dimensions['B'].width = 35.0
        rr = 2
        total = 0
        for name in sorted(data.keys()):
            setcell(ws, rr, 1, data[name])
            setcell(ws, rr, 2, name.upper())
            total += data[name]
            rr += 1
        rr += 1
        setcell(ws, rr, 1, 'TOTAL', bold=True)
        setcell(ws, rr, 2, total, bold=True)

    write_agg_sheet("Picking del dia", picking)
    write_agg_sheet("Pedido a taller", taller)

    # Save output
    wb.save(out_path)

    # Actualizar CSV de control con nuevas ventas vistas HOY
    fecha_hoy = datetime.date.today().isoformat()
    append_control(control_csv, nuevas_vids, fecha_hoy)

    print("OK ->", out_path)
    print("Ventas nuevas hoy:", len(nuevas_vids))

# -----------------------
# CLI
# -----------------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: build_ml_report.py src.xlsx out.xlsx [control_csv]")
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2]
    control = sys.argv[3] if len(sys.argv) > 3 else None
    build_report(src, out, control)
